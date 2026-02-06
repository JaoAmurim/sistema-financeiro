#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Sistema de Acompanhamento de Investimentos
Especializado para carteira de FIIs e Ações
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import PieChart, LineChart, Reference
from datetime import datetime, timedelta

def criar_sistema_investimentos():
    """Cria planilha completa de controle de investimentos"""
    wb = Workbook()
    wb.remove(wb.active)  # Remove a planilha padrão
    
    # ========== DEFINIÇÕES DE ESTILO ==========
    # Cores profissionais
    cor_header_azul = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    cor_header_verde = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    cor_header_laranja = PatternFill(start_color="E65100", end_color="E65100", fill_type="solid")
    cor_header_roxo = PatternFill(start_color="6A1B9A", end_color="6A1B9A", fill_type="solid")
    cor_header_vermelho = PatternFill(start_color="C62828", end_color="C62828", fill_type="solid")
    
    cor_positivo = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    cor_negativo = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    cor_neutro = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    cor_info = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
    cor_destaque = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")
    
    # Fontes
    font_header = Font(bold=True, size=11, color="FFFFFF")
    font_titulo = Font(bold=True, size=14, color="1F4E78")
    
    # Bordas
    border_thin = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    border_thick = Border(
        left=Side(style='medium'),
        right=Side(style='medium'),
        top=Side(style='medium'),
        bottom=Side(style='medium')
    )
    
    # ========== ABA 1: DASHBOARD ==========
    ws_dash = wb.create_sheet("📊 Dashboard")
    
    # Título
    ws_dash['B2'] = "🎯 MEU SISTEMA DE INVESTIMENTOS"
    ws_dash['B2'].font = Font(bold=True, size=18, color="1F4E78")
    ws_dash['B2'].alignment = Alignment(horizontal='center')
    ws_dash.merge_cells('B2:H2')
    ws_dash.row_dimensions[2].height = 30
    
    # Subtítulo com data
    ws_dash['B3'] = "📅 Última Atualização:"
    ws_dash['B3'].font = Font(size=10, italic=True)
    ws_dash['C3'] = '=HOJE()'
    ws_dash['C3'].number_format = 'DD/MM/YYYY'
    ws_dash['C3'].font = Font(size=10, bold=True)
    
    # Cards de resumo
    ws_dash['B5'] = "💰 PATRIMÔNIO ATUAL"
    ws_dash['B5'].font = Font(bold=True, size=13, color="FFFFFF")
    ws_dash['B5'].fill = cor_header_verde
    ws_dash['B5'].alignment = Alignment(horizontal='center')
    ws_dash.merge_cells('B5:C5')
    
    ws_dash['B6'] = '=SOMARPRODUTO(Carteira!D6:D20,Carteira!E6:E20)'
    ws_dash['B6'].number_format = 'R$ #,##0.00'
    ws_dash['B6'].font = Font(bold=True, size=20, color="2E7D32")
    ws_dash['B6'].alignment = Alignment(horizontal='center')
    ws_dash['B6'].fill = cor_positivo
    ws_dash['B6'].border = border_thick
    ws_dash.merge_cells('B6:C6')
    ws_dash.row_dimensions[6].height = 40
    
    ws_dash['E5'] = "📈 RENTABILIDADE 2026"
    ws_dash['E5'].font = Font(bold=True, size=13, color="FFFFFF")
    ws_dash['E5'].fill = cor_header_azul
    ws_dash['E5'].alignment = Alignment(horizontal='center')
    ws_dash.merge_cells('E5:F5')
    
    ws_dash['E6'] = '=((B6-Performance!C6)/Performance!C6)'
    ws_dash['E6'].number_format = '0.00%'
    ws_dash['E6'].font = Font(bold=True, size=20)
    ws_dash['E6'].alignment = Alignment(horizontal='center')
    ws_dash['E6'].fill = cor_info
    ws_dash['E6'].border = border_thick
    ws_dash.merge_cells('E6:F6')
    
    ws_dash['G5'] = "💵 PROVENTOS/MÊS"
    ws_dash['G5'].font = Font(bold=True, size=13, color="FFFFFF")
    ws_dash['G5'].fill = cor_header_verde
    ws_dash['G5'].alignment = Alignment(horizontal='center')
    ws_dash.merge_cells('G5:H5')
    
    ws_dash['G6'] = '=SOMASE(Proventos!B:B,">="&DATA(ANO(HOJE()),MÊS(HOJE()),1),Proventos!E:E)'
    ws_dash['G6'].number_format = 'R$ #,##0.00'
    ws_dash['G6'].font = Font(bold=True, size=20, color="2E7D32")
    ws_dash['G6'].alignment = Alignment(horizontal='center')
    ws_dash['G6'].fill = cor_positivo
    ws_dash['G6'].border = border_thick
    ws_dash.merge_cells('G6:H6')
    
    # Performance vs CDI
    ws_dash['B9'] = "📊 PERFORMANCE vs BENCHMARKS"
    ws_dash['B9'].font = Font(bold=True, size=12, color="FFFFFF")
    ws_dash['B9'].fill = cor_header_azul
    ws_dash['B9'].alignment = Alignment(horizontal='center')
    ws_dash.merge_cells('B9:H9')
    
    headers_perf = ['B10', 'C10', 'D10', 'E10', 'F10']
    valores_perf = ['Indicador', 'Rentabilidade', 'Vs. Minha Carteira', '% do CDI', 'Status']
    for cell, valor in zip(headers_perf, valores_perf):
        ws_dash[cell] = valor
        ws_dash[cell].fill = cor_header_azul
        ws_dash[cell].font = font_header
        ws_dash[cell].alignment = Alignment(horizontal='center')
        ws_dash[cell].border = border_thin
    
    ws_dash['B11'] = 'Minha Carteira'
    ws_dash['C11'] = '=E6'
    ws_dash['C11'].number_format = '0.00%'
    ws_dash['D11'] = '-'
    ws_dash['E11'] = '=(E6/Performance!C8)'
    ws_dash['E11'].number_format = '0.00%'
    ws_dash['F11'] = '=SE(E11>=1.1,"🔥 Excelente!",SE(E11>=1,"✅ Batendo!","⚠️ Abaixo"))'
    
    ws_dash['B12'] = 'CDI'
    ws_dash['C12'] = '=Performance!C8'
    ws_dash['C12'].number_format = '0.00%'
    ws_dash['D12'] = '=E6-C12'
    ws_dash['D12'].number_format = '+0.00%;-0.00%'
    ws_dash['E12'] = '100,00%'
    ws_dash['F12'] = 'Benchmark'
    
    for row in [11, 12]:
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws_dash[f'{col}{row}'].border = border_thin
            ws_dash[f'{col}{row}'].alignment = Alignment(horizontal='center')
    
    # Composição da carteira
    ws_dash['B15'] = "📂 COMPOSIÇÃO DA CARTEIRA"
    ws_dash['B15'].font = Font(bold=True, size=12, color="FFFFFF")
    ws_dash['B15'].fill = cor_header_roxo
    ws_dash['B15'].alignment = Alignment(horizontal='center')
    ws_dash.merge_cells('B15:F15')
    
    headers_comp = ['B16', 'C16', 'D16', 'E16', 'F16']
    valores_comp = ['Ativo', 'Cotas', 'Valor Investido', 'Valor Atual', '% Carteira']
    for cell, valor in zip(headers_comp, valores_comp):
        ws_dash[cell] = valor
        ws_dash[cell].fill = cor_header_roxo
        ws_dash[cell].font = font_header
        ws_dash[cell].alignment = Alignment(horizontal='center')
        ws_dash[cell].border = border_thin
    
    # Fórmulas para buscar dados da carteira
    for idx, row in enumerate([17, 18, 19], start=1):
        ws_dash[f'B{row}'] = f'=ÍNDICE(Carteira!B:B,{idx+5})'
        ws_dash[f'C{row}'] = f'=ÍNDICE(Carteira!D:D,{idx+5})'
        ws_dash[f'D{row}'] = f'=ÍNDICE(Carteira!F:F,{idx+5})'
        ws_dash[f'D{row}'].number_format = 'R$ #,##0.00'
        ws_dash[f'E{row}'] = f'=C{row}*ÍNDICE(Carteira!E:E,{idx+5})'
        ws_dash[f'E{row}'].number_format = 'R$ #,##0.00'
        ws_dash[f'F{row}'] = f'=E{row}/B6'
        ws_dash[f'F{row}'].number_format = '0.0%'
        
        for col in ['B', 'C', 'D', 'E', 'F']:
            ws_dash[f'{col}{row}'].border = border_thin
            ws_dash[f'{col}{row}'].alignment = Alignment(horizontal='center')
    
    # Meta do mês
    ws_dash['B22'] = "🎯 META DO MÊS ATUAL"
    ws_dash['B22'].font = Font(bold=True, size=12, color="FFFFFF")
    ws_dash['B22'].fill = cor_header_laranja
    ws_dash['B22'].alignment = Alignment(horizontal='center')
    ws_dash.merge_cells('B22:F22')
    
    ws_dash['B23'] = 'Meta de Aporte:'
    ws_dash['C23'] = '=ÍNDICE(Aportes!D:D,MÊS(HOJE())+5)'
    ws_dash['C23'].number_format = 'R$ #,##0.00'
    ws_dash['C23'].font = Font(bold=True, color="E65100")
    
    ws_dash['B24'] = 'Já Aportado:'
    ws_dash['C24'] = '=ÍNDICE(Aportes!E:E,MÊS(HOJE())+5)'
    ws_dash['C24'].number_format = 'R$ #,##0.00'
    ws_dash['C24'].font = Font(bold=True, color="2E7D32")
    
    ws_dash['B25'] = 'Falta Aportar:'
    ws_dash['C25'] = '=C23-C24'
    ws_dash['C25'].number_format = 'R$ #,##0.00'
    ws_dash['C25'].font = Font(bold=True)
    ws_dash['C25'].fill = cor_neutro
    
    # Larguras
    ws_dash.column_dimensions['B'].width = 20
    ws_dash.column_dimensions['C'].width = 15
    ws_dash.column_dimensions['D'].width = 15
    ws_dash.column_dimensions['E'].width = 18
    ws_dash.column_dimensions['F'].width = 18
    ws_dash.column_dimensions['G'].width = 18
    ws_dash.column_dimensions['H'].width = 18
    
    # ========== ABA 2: CARTEIRA ==========
    ws_carteira = wb.create_sheet("💼 Carteira")
    
    ws_carteira['B2'] = "💼 MINHA CARTEIRA DE INVESTIMENTOS"
    ws_carteira['B2'].font = Font(bold=True, size=16, color="FFFFFF")
    ws_carteira['B2'].fill = cor_header_roxo
    ws_carteira['B2'].alignment = Alignment(horizontal='center')
    ws_carteira.merge_cells('B2:I2')
    ws_carteira.row_dimensions[2].height = 35
    
    ws_carteira['B3'] = "💡 INSTRUÇÕES: Adicione seus ativos abaixo. Atualize as cotações na coluna E para ver o valor atual automaticamente."
    ws_carteira['B3'].font = Font(italic=True, size=9, color="666666")
    ws_carteira.merge_cells('B3:I3')
    
    headers_cart = ['B5', 'C5', 'D5', 'E5', 'F5', 'G5', 'H5', 'I5']
    valores_cart = ['🏢 Ativo', '📂 Tipo', '🔢 Cotas', '💵 Cotação Atual', '💰 Preço Médio', '📊 Total Investido', '📈 Valor Atual', '📊 Rent. %']
    cores_cart = [cor_header_azul, cor_header_roxo, cor_header_verde, cor_header_laranja, cor_header_laranja, cor_header_azul, cor_header_verde, cor_header_roxo]
    
    for cell, valor, cor in zip(headers_cart, valores_cart, cores_cart):
        ws_carteira[cell] = valor
        ws_carteira[cell].fill = cor
        ws_carteira[cell].font = font_header
        ws_carteira[cell].alignment = Alignment(horizontal='center', wrap_text=True)
        ws_carteira[cell].border = border_thick
    ws_carteira.row_dimensions[5].height = 35
    
    # Primeira linha com instruções
    ws_carteira['B6'] = '👉 Digite aqui (ex: MXRF11)'
    ws_carteira['B6'].font = Font(italic=True, size=9, color="1F4E78")
    ws_carteira['C6'] = 'FII'
    ws_carteira['C6'].font = Font(italic=True, size=9)
    
    # Validação de dados para tipo
    dv_tipo = DataValidation(type="list", formula1='"FII,Ação,Renda Fixa"', allow_blank=False)
    dv_tipo.prompt = 'Selecione: FII, Ação ou Renda Fixa'
    dv_tipo.promptTitle = 'Tipo de Ativo'
    ws_carteira.add_data_validation(dv_tipo)
    dv_tipo.add('C6:C20')
    
    # Adicionar fórmulas nas colunas
    ws_carteira['F6'] = '=D6*E6'
    ws_carteira['F6'].number_format = 'R$ #,##0.00'
    ws_carteira['G6'] = '=D6*E6'
    ws_carteira['G6'].number_format = 'R$ #,##0.00'
    ws_carteira['H6'] = '=(G6-F6)/F6'
    ws_carteira['H6'].number_format = '0.00%'
    
    # Formatação condicional para rentabilidade
    ws_carteira.conditional_formatting.add('H6:H20',
        CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=True, fill=cor_positivo, font=Font(color="2E7D32", bold=True)))
    ws_carteira.conditional_formatting.add('H6:H20',
        CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, fill=cor_negativo, font=Font(color="C62828", bold=True)))
    
    # Totais
    ws_carteira['B22'] = "💰 TOTAIS"
    ws_carteira['B22'].font = Font(bold=True, size=12, color="FFFFFF")
    ws_carteira['B22'].fill = cor_header_azul
    ws_carteira['B22'].alignment = Alignment(horizontal='center')
    ws_carteira.merge_cells('B22:E22')
    
    ws_carteira['F22'] = '=SOMA(F6:F20)'
    ws_carteira['F22'].number_format = 'R$ #,##0.00'
    ws_carteira['F22'].font = Font(bold=True, size=12)
    ws_carteira['F22'].fill = cor_info
    ws_carteira['F22'].border = border_thick
    
    ws_carteira['G22'] = '=SOMA(G6:G20)'
    ws_carteira['G22'].number_format = 'R$ #,##0.00'
    ws_carteira['G22'].font = Font(bold=True, size=12)
    ws_carteira['G22'].fill = cor_positivo
    ws_carteira['G22'].border = border_thick
    
    ws_carteira['H22'] = '=(G22-F22)/F22'
    ws_carteira['H22'].number_format = '0.00%'
    ws_carteira['H22'].font = Font(bold=True, size=12)
    ws_carteira['H22'].fill = cor_destaque
    ws_carteira['H22'].border = border_thick
    
    # Larguras
    ws_carteira.column_dimensions['B'].width = 18
    ws_carteira.column_dimensions['C'].width = 12
    ws_carteira.column_dimensions['D'].width = 10
    ws_carteira.column_dimensions['E'].width = 15
    ws_carteira.column_dimensions['F'].width = 15
    ws_carteira.column_dimensions['G'].width = 15
    ws_carteira.column_dimensions['H'].width = 15
    ws_carteira.column_dimensions['I'].width = 12
    
    # ========== ABA 3: APORTES ==========
    ws_aportes = wb.create_sheet("📅 Aportes")
    
    ws_aportes['B2'] = "📅 PLANEJAMENTO DE APORTES MENSAIS"
    ws_aportes['B2'].font = Font(bold=True, size=16, color="FFFFFF")
    ws_aportes['B2'].fill = cor_header_laranja
    ws_aportes['B2'].alignment = Alignment(horizontal='center')
    ws_aportes.merge_cells('B2:G2')
    ws_aportes.row_dimensions[2].height = 35
    
    ws_aportes['B3'] = "🎯 OBJETIVO: Acompanhe suas metas de aporte mensal e quanto já foi investido"
    ws_aportes['B3'].font = Font(italic=True, size=9, color="666666")
    ws_aportes.merge_cells('B3:G3')
    
    headers_ap = ['B5', 'C5', 'D5', 'E5', 'F5', 'G5']
    valores_ap = ['📅 Mês', '🏢 Meta de Cotas', '💰 Meta Valor (R$)', '✅ Já Aportado', '📊 % Realizado', '📝 Status']
    cores_ap = [cor_header_azul, cor_header_verde, cor_header_laranja, cor_header_verde, cor_header_roxo, cor_header_azul]
    
    for cell, valor, cor in zip(headers_ap, valores_ap, cores_ap):
        ws_aportes[cell] = valor
        ws_aportes[cell].fill = cor
        ws_aportes[cell].font = font_header
        ws_aportes[cell].alignment = Alignment(horizontal='center', wrap_text=True)
        ws_aportes[cell].border = border_thick
    ws_aportes.row_dimensions[5].height = 35
    
    # Meses do ano
    meses = ['Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 
             'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
    
    for idx, mes in enumerate(meses, start=6):
        ws_aportes[f'B{idx}'] = mes
        ws_aportes[f'B{idx}'].font = Font(bold=True)
        ws_aportes[f'B{idx}'].border = border_thin
        
        # Fórmulas
        ws_aportes[f'F{idx}'] = f'=SE(E{idx}=0,"⏳ Aguardando",SE(E{idx}>=D{idx},"✅ Meta Atingida!","📊 Em Progresso"))'
        ws_aportes[f'F{idx}'].border = border_thin
        ws_aportes[f'F{idx}'].alignment = Alignment(horizontal='center')
        
        ws_aportes[f'G{idx}'] = f'=E{idx}/D{idx}'
        ws_aportes[f'G{idx}'].number_format = '0%'
        ws_aportes[f'G{idx}'].border = border_thin
        ws_aportes[f'G{idx}'].alignment = Alignment(horizontal='center')
        
        for col in ['C', 'D', 'E']:
            ws_aportes[f'{col}{idx}'].border = border_thin
            ws_aportes[f'{col}{idx}'].alignment = Alignment(horizontal='center')
            if col in ['D', 'E']:
                ws_aportes[f'{col}{idx}'].number_format = 'R$ #,##0.00'
    
    # Formatação condicional
    ws_aportes.conditional_formatting.add('G6:G17',
        CellIsRule(operator='greaterThanOrEqual', formula=['1'], stopIfTrue=True, fill=cor_positivo))
    ws_aportes.conditional_formatting.add('G6:G17',
        CellIsRule(operator='between', formula=['0.5', '0.99'], stopIfTrue=True, fill=cor_neutro))
    ws_aportes.conditional_formatting.add('G6:G17',
        CellIsRule(operator='lessThan', formula=['0.5'], stopIfTrue=True, fill=cor_negativo))
    
    # Totais anuais
    ws_aportes['B19'] = "📊 TOTAL ANUAL"
    ws_aportes['B19'].font = Font(bold=True, size=11, color="FFFFFF")
    ws_aportes['B19'].fill = cor_header_azul
    ws_aportes['B19'].alignment = Alignment(horizontal='center')
    ws_aportes.merge_cells('B19:C19')
    
    ws_aportes['D19'] = '=SOMA(D6:D17)'
    ws_aportes['D19'].number_format = 'R$ #,##0.00'
    ws_aportes['D19'].font = Font(bold=True, size=12)
    ws_aportes['D19'].fill = cor_info
    ws_aportes['D19'].border = border_thick
    
    ws_aportes['E19'] = '=SOMA(E6:E17)'
    ws_aportes['E19'].number_format = 'R$ #,##0.00'
    ws_aportes['E19'].font = Font(bold=True, size=12)
    ws_aportes['E19'].fill = cor_positivo
    ws_aportes['E19'].border = border_thick
    
    ws_aportes['G19'] = '=E19/D19'
    ws_aportes['G19'].number_format = '0%'
    ws_aportes['G19'].font = Font(bold=True, size=12)
    ws_aportes['G19'].fill = cor_destaque
    ws_aportes['G19'].border = border_thick
    
    # Larguras
    ws_aportes.column_dimensions['B'].width = 15
    ws_aportes.column_dimensions['C'].width = 15
    ws_aportes.column_dimensions['D'].width = 18
    ws_aportes.column_dimensions['E'].width = 18
    ws_aportes.column_dimensions['F'].width = 20
    ws_aportes.column_dimensions['G'].width = 15
    
    # ========== ABA 4: PROVENTOS ==========
    ws_prov = wb.create_sheet("💰 Proventos")
    
    ws_prov['B2'] = "💰 CONTROLE DE PROVENTOS RECEBIDOS"
    ws_prov['B2'].font = Font(bold=True, size=16, color="FFFFFF")
    ws_prov['B2'].fill = cor_header_verde
    ws_prov['B2'].alignment = Alignment(horizontal='center')
    ws_prov.merge_cells('B2:F2')
    ws_prov.row_dimensions[2].height = 35
    
    ws_prov['B3'] = "📝 Registre aqui todos os dividendos, JCP e rendimentos que você receber"
    ws_prov['B3'].font = Font(italic=True, size=9, color="666666")
    ws_prov.merge_cells('B3:F3')
    
    headers_pr = ['B5', 'C5', 'D5', 'E5', 'F5']
    valores_pr = ['📅 Data', '🏢 Ativo', '📂 Tipo', '💵 Valor/Cota', '💰 Total Recebido']
    cores_pr = [cor_header_azul, cor_header_azul, cor_header_verde, cor_header_laranja, cor_header_verde]
    
    for cell, valor, cor in zip(headers_pr, valores_pr, cores_pr):
        ws_prov[cell] = valor
        ws_prov[cell].fill = cor
        ws_prov[cell].font = font_header
        ws_prov[cell].alignment = Alignment(horizontal='center', wrap_text=True)
        ws_prov[cell].border = border_thick
    ws_prov.row_dimensions[5].height = 30
    
    # Validação
    dv_tipo_prov = DataValidation(type="list", formula1='"Dividendo,JCP,Rendimento,Juros"', allow_blank=False)
    dv_tipo_prov.prompt = 'Selecione o tipo de provento'
    dv_tipo_prov.promptTitle = 'Tipo'
    ws_prov.add_data_validation(dv_tipo_prov)
    dv_tipo_prov.add('D6:D100')
    
    # Primeira linha
    ws_prov['B6'] = '=HOJE()'
    ws_prov['B6'].number_format = 'DD/MM/YYYY'
    ws_prov['B6'].font = Font(italic=True, size=9)
    
    ws_prov['C6'] = '👉 Digite o código do ativo'
    ws_prov['C6'].font = Font(italic=True, size=9, color="1F4E78")
    
    # Resumo mensal
    ws_prov['B10'] = "📊 RESUMO POR MÊS"
    ws_prov['B10'].font = Font(bold=True, size=12, color="FFFFFF")
    ws_prov['B10'].fill = cor_header_azul
    ws_prov['B10'].alignment = Alignment(horizontal='center')
    ws_prov.merge_cells('B10:D10')
    
    ws_prov['B11'] = 'Mês Atual:'
    ws_prov['C11'] = '=SOMASE(B:B,">="&DATA(ANO(HOJE()),MÊS(HOJE()),1),E:E)'
    ws_prov['C11'].number_format = 'R$ #,##0.00'
    ws_prov['C11'].font = Font(bold=True, size=12, color="2E7D32")
    ws_prov['C11'].fill = cor_positivo
    
    ws_prov['B12'] = 'Média Mensal:'
    ws_prov['C12'] = '=SOMA(E:E)/MÊS(HOJE())'
    ws_prov['C12'].number_format = 'R$ #,##0.00'
    ws_prov['C12'].font = Font(bold=True)
    ws_prov['C12'].fill = cor_info
    
    ws_prov['B13'] = 'Projeção Anual:'
    ws_prov['C13'] = '=C12*12'
    ws_prov['C13'].number_format = 'R$ #,##0.00'
    ws_prov['C13'].font = Font(bold=True, size=12)
    ws_prov['C13'].fill = cor_destaque
    
    ws_prov['B14'] = 'Acumulado 2026:'
    ws_prov['C14'] = '=SOMA(E:E)'
    ws_prov['C14'].number_format = 'R$ #,##0.00'
    ws_prov['C14'].font = Font(bold=True, size=13, color="2E7D32")
    ws_prov['C14'].fill = cor_positivo
    ws_prov['C14'].border = border_thick
    
    # Larguras
    ws_prov.column_dimensions['B'].width = 15
    ws_prov.column_dimensions['C'].width = 18
    ws_prov.column_dimensions['D'].width = 15
    ws_prov.column_dimensions['E'].width = 15
    ws_prov.column_dimensions['F'].width = 18
    
    # ========== ABA 5: PERFORMANCE ==========
    ws_perf = wb.create_sheet("📈 Performance")
    
    ws_perf['B2'] = "📈 ACOMPANHAMENTO DE PERFORMANCE"
    ws_perf['B2'].font = Font(bold=True, size=16, color="FFFFFF")
    ws_perf['B2'].fill = cor_header_azul
    ws_perf['B2'].alignment = Alignment(horizontal='center')
    ws_perf.merge_cells('B2:F2')
    ws_perf.row_dimensions[2].height = 35
    
    ws_perf['B3'] = "📊 Acompanhe a evolução do seu patrimônio mês a mês"
    ws_perf['B3'].font = Font(italic=True, size=9, color="666666")
    ws_perf.merge_cells('B3:F3')
    
    headers_pf = ['B5', 'C5', 'D5', 'E5', 'F5']
    valores_pf = ['📅 Mês', '💰 Patrimônio', '📊 Variação %', '📈 CDI Acum.', '🎯 Vs. CDI']
    cores_pf = [cor_header_azul, cor_header_verde, cor_header_roxo, cor_header_laranja, cor_header_azul]
    
    for cell, valor, cor in zip(headers_pf, valores_pf, cores_pf):
        ws_perf[cell] = valor
        ws_perf[cell].fill = cor
        ws_perf[cell].font = font_header
        ws_perf[cell].alignment = Alignment(horizontal='center', wrap_text=True)
        ws_perf[cell].border = border_thick
    ws_perf.row_dimensions[5].height = 30
    
    # Linha inicial (Janeiro)
    ws_perf['B6'] = '2026-01'
    ws_perf['B6'].font = Font(bold=True)
    ws_perf['C6'] = 0  # Patrimônio inicial - usuário preenche
    ws_perf['C6'].number_format = 'R$ #,##0.00'
    
    # Fevereiro em diante (usuário preenche conforme os meses passam)
    for idx, mes in enumerate(['02', '03', '04', '05', '06', '07', '08', '09', '10', '11', '12'], start=7):
        ws_perf[f'B{idx}'] = f'2026-{mes}'
        ws_perf[f'B{idx}'].font = Font(bold=True)
        ws_perf[f'B{idx}'].border = border_thin
        
        ws_perf[f'C{idx}'].number_format = 'R$ #,##0.00'
        ws_perf[f'C{idx}'].border = border_thin
        
        ws_perf[f'D{idx}'] = f'=(C{idx}-C{idx-1})/C{idx-1}'
        ws_perf[f'D{idx}'].number_format = '0.00%'
        ws_perf[f'D{idx}'].border = border_thin
        
        ws_perf[f'E{idx}'].number_format = '0.00%'
        ws_perf[f'E{idx}'].border = border_thin
        
        ws_perf[f'F{idx}'] = f'=D{idx}-E{idx}'
        ws_perf[f'F{idx}'].number_format = '+0.00%;-0.00%'
        ws_perf[f'F{idx}'].border = border_thin
    
    # Campo para CDI acumulado do ano
    ws_perf['B8'] = 'CDI Acumulado 2026:'
    ws_perf['B8'].font = Font(bold=True)
    ws_perf['C8'] = 0  # Usuário atualiza com CDI real
    ws_perf['C8'].number_format = '0.00%'
    ws_perf['C8'].fill = cor_info
    ws_perf['C8'].font = Font(bold=True)
    
    # Formatação condicional
    ws_perf.conditional_formatting.add('D7:D17',
        CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=True, fill=cor_positivo, font=Font(color="2E7D32", bold=True)))
    ws_perf.conditional_formatting.add('D7:D17',
        CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, fill=cor_negativo, font=Font(color="C62828", bold=True)))
    
    ws_perf.conditional_formatting.add('F6:F17',
        CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=True, fill=cor_positivo))
    ws_perf.conditional_formatting.add('F6:F17',
        CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=True, fill=cor_negativo))
    
    # Larguras
    ws_perf.column_dimensions['B'].width = 15
    ws_perf.column_dimensions['C'].width = 18
    ws_perf.column_dimensions['D'].width = 15
    ws_perf.column_dimensions['E'].width = 15
    ws_perf.column_dimensions['F'].width = 15
    
    # ========== ABA 6: METAS ==========
    ws_metas = wb.create_sheet("🎯 Metas")
    
    ws_metas['B2'] = "🎯 MINHAS METAS DE LONGO PRAZO"
    ws_metas['B2'].font = Font(bold=True, size=16, color="FFFFFF")
    ws_metas['B2'].fill = cor_header_laranja
    ws_metas['B2'].alignment = Alignment(horizontal='center')
    ws_metas.merge_cells('B2:F2')
    ws_metas.row_dimensions[2].height = 35
    
    # Meta de Patrimônio
    ws_metas['B5'] = "💰 META DE PATRIMÔNIO"
    ws_metas['B5'].font = Font(bold=True, size=12, color="FFFFFF")
    ws_metas['B5'].fill = cor_header_verde
    ws_metas['B5'].alignment = Alignment(horizontal='center')
    ws_metas.merge_cells('B5:D5')
    
    ws_metas['B6'] = 'Meta para 2026:'
    ws_metas['C6'] = 0  # Usuário define
    ws_metas['C6'].number_format = 'R$ #,##0.00'
    ws_metas['C6'].font = Font(bold=True, size=12)
    ws_metas['C6'].fill = cor_destaque
    
    ws_metas['B7'] = 'Patrimônio Atual:'
    ws_metas['C7'] = '=Dashboard!B6'
    ws_metas['C7'].number_format = 'R$ #,##0.00'
    ws_metas['C7'].font = Font(bold=True, size=12, color="2E7D32")
    ws_metas['C7'].fill = cor_positivo
    
    ws_metas['B8'] = '% Atingido:'
    ws_metas['C8'] = '=C7/C6'
    ws_metas['C8'].number_format = '0.0%'
    ws_metas['C8'].font = Font(bold=True, size=14)
    ws_metas['C8'].fill = cor_info
    ws_metas['C8'].border = border_thick
    
    ws_metas['B9'] = 'Falta para a Meta:'
    ws_metas['C9'] = '=C6-C7'
    ws_metas['C9'].number_format = 'R$ #,##0.00'
    ws_metas['C9'].font = Font(bold=True)
    
    # Meta de Renda Passiva
    ws_metas['B12'] = "💵 META DE RENDA PASSIVA"
    ws_metas['B12'].font = Font(bold=True, size=12, color="FFFFFF")
    ws_metas['B12'].fill = cor_header_verde
    ws_metas['B12'].alignment = Alignment(horizontal='center')
    ws_metas.merge_cells('B12:D12')
    
    ws_metas['B13'] = 'Meta Mensal:'
    ws_metas['C13'] = 0  # Usuário define (ex: R$ 1.000)
    ws_metas['C13'].number_format = 'R$ #,##0.00'
    ws_metas['C13'].font = Font(bold=True, size=12)
    ws_metas['C13'].fill = cor_destaque
    
    ws_metas['B14'] = 'Recebendo Agora:'
    ws_metas['C14'] = '=Proventos!C12'
    ws_metas['C14'].number_format = 'R$ #,##0.00'
    ws_metas['C14'].font = Font(bold=True, size=12, color="2E7D32")
    ws_metas['C14'].fill = cor_positivo
    
    ws_metas['B15'] = '% Atingido:'
    ws_metas['C15'] = '=C14/C13'
    ws_metas['C15'].number_format = '0.0%'
    ws_metas['C15'].font = Font(bold=True, size=14)
    ws_metas['C15'].fill = cor_info
    ws_metas['C15'].border = border_thick
    
    ws_metas['B16'] = 'Falta para a Meta:'
    ws_metas['C16'] = '=C13-C14'
    ws_metas['C16'].number_format = 'R$ #,##0.00'
    ws_metas['C16'].font = Font(bold=True)
    
    # Projeções
    ws_metas['B19'] = "📊 PROJEÇÕES"
    ws_metas['B19'].font = Font(bold=True, size=12, color="FFFFFF")
    ws_metas['B19'].fill = cor_header_azul
    ws_metas['B19'].alignment = Alignment(horizontal='center')
    ws_metas.merge_cells('B19:D19')
    
    ws_metas['B20'] = 'Mantendo ritmo atual:'
    ws_metas['B20'].font = Font(bold=True)
    
    ws_metas['B21'] = 'Patrimônio em 1 ano:'
    ws_metas['C21'] = '=Dashboard!B6*(1+Performance!C8)'
    ws_metas['C21'].number_format = 'R$ #,##0.00'
    
    ws_metas['B22'] = 'Renda passiva em 1 ano:'
    ws_metas['C22'] = '=Proventos!C12*1.1'
    ws_metas['C22'].number_format = 'R$ #,##0.00'
    
    # Larguras
    ws_metas.column_dimensions['B'].width = 25
    ws_metas.column_dimensions['C'].width = 20
    ws_metas.column_dimensions['D'].width = 15
    
    # ========== SALVAR ARQUIVO ==========
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")
    filename = f"Sistema_Investimentos_{timestamp}.xlsx"
    wb.save(filename)
    
    print("=" * 80)
    print("✅ SISTEMA DE INVESTIMENTOS CRIADO COM SUCESSO!")
    print("=" * 80)
    print(f"📁 Arquivo: {filename}")
    print(f"📊 Total de abas: {len(wb.sheetnames)}")
    print("\n🎯 ABAS CRIADAS:")
    print("   1. 📊 Dashboard - Visão geral do patrimônio e performance")
    print("   2. 💼 Carteira - Controle de suas cotas e ativos")
    print("   3. 📅 Aportes - Planejamento mensal de investimentos")
    print("   4. 💰 Proventos - Registro de dividendos recebidos")
    print("   5. 📈 Performance - Evolução patrimonial mensal")
    print("   6. 🎯 Metas - Objetivos de patrimônio e renda passiva")
    print("\n💡 PRÓXIMOS PASSOS:")
    print("   1. Abra o arquivo no Excel/LibreOffice")
    print("   2. Na aba 'Carteira', adicione seus ativos (MXRF11, GGRC11, ITSA3)")
    print("   3. Preencha quantas cotas você tem e o preço médio de compra")
    print("   4. Atualize as cotações atuais regularmente")
    print("   5. Na aba 'Aportes', defina suas metas mensais")
    print("   6. Registre os proventos conforme receber")
    print("   7. Acompanhe seu progresso no Dashboard!")
    print("\n🚀 Bons investimentos!")
    print("=" * 80)

if __name__ == "__main__":
    criar_sistema_investimentos()
